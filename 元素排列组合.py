import itertools
from openpyxl import load_workbook
import os


def createpermutations(string):
    result = list(itertools.permutations(string, len(string)))
    # print(result)
    result1 = ["".join(x) for x in result]
    # print(result1)
    return result1


def main():
    keyIn = input("1、输入字母按回车\n2、直接按回车用表格\n")
    if keyIn == "":
        filepath = input("\n文件路径：\n")
        filepath = filepath.replace("\"", "").replace("\'", "")
        wb = load_workbook(filepath)
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]  # index为0为第一张表
        path = os.path.splitext(filepath)[0] + ".csv"
        print(path)
    else:
        # elements = input("输入字母, 回车结束")
        result = createpermutations(keyIn)
        print(result)
        print(str(result).replace('\'', '').replace('[', '').replace(']', '').replace(',', ''))
        input("搞定,按回车结束")
        return
    text = []
    for i in range(1, ws.max_row + 1):
        text.append(ws.cell(i, 1).value)
    fp = open(path, 'a')
    for i in range(len(text)):
        fp.write(text[i] + ',')
        result = createpermutations(text[i])
        fp.write(",".join(result) + '\n')
    fp.close()
    input("搞定,按回车结束")


if __name__ == '__main__':
    main()
